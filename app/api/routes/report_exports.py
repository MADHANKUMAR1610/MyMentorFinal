from io import BytesIO

from fastapi import APIRouter, Depends
from fastapi.responses import StreamingResponse
from sqlalchemy.ext.asyncio import AsyncSession
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from app.database.database import get_db
from app.api.dependencies import get_current_company_admin
from app.services.analytics_service import OrganizationAnalyticsService


router = APIRouter(
    prefix="/reports",
    tags=["Reports Export"],
)


# ============================================================
# EXCEL STYLING
# ============================================================

HEADER_FILL = PatternFill(
    fill_type="solid",
    fgColor="1F4E78",
)

HEADER_FONT = Font(
    bold=True,
    color="FFFFFF",
)

TITLE_FONT = Font(
    bold=True,
    size=16,
)

SUBTITLE_FONT = Font(
    bold=True,
    size=11,
)

THIN_BORDER = Border(
    left=Side(style="thin", color="D9E1F2"),
    right=Side(style="thin", color="D9E1F2"),
    top=Side(style="thin", color="D9E1F2"),
    bottom=Side(style="thin", color="D9E1F2"),
)


def style_header(ws, row_number: int):
    """
    Apply header styling to a worksheet row.
    """

    for cell in ws[row_number]:
        if cell.value is not None:
            cell.fill = HEADER_FILL
            cell.font = HEADER_FONT
            cell.alignment = Alignment(
                horizontal="center",
                vertical="center",
            )
            cell.border = THIN_BORDER


def auto_width(ws):
    """
    Automatically adjust Excel column widths.
    """

    for column_cells in ws.columns:

        max_length = 0
        column_letter = get_column_letter(
            column_cells[0].column
        )

        for cell in column_cells:

            if cell.value is not None:

                value_length = len(
                    str(cell.value)
                )

                if value_length > max_length:
                    max_length = value_length

        ws.column_dimensions[
            column_letter
        ].width = min(
            max(max_length + 3, 12),
            45,
        )


def workbook_to_response(
    workbook: Workbook,
    filename: str,
):

    output = BytesIO()

    workbook.save(output)

    output.seek(0)

    return StreamingResponse(
        output,
        media_type=(
            "application/vnd.openxmlformats-officedocument."
            "spreadsheetml.sheet"
        ),
        headers={
            "Content-Disposition": (
                f'attachment; filename="{filename}"'
            )
        },
    )


# ============================================================
# 1. EXPORT OVERVIEW
# ============================================================

@router.get(
    "/export/overview",
)
async def export_overview(
    current_user=Depends(
        get_current_company_admin
    ),
    db: AsyncSession = Depends(get_db),
):

    # --------------------------------------------------------
    # COMPANY
    # --------------------------------------------------------

    company_id = current_user.company_id

    if company_id is None:
        from fastapi import HTTPException

        raise HTTPException(
            status_code=400,
            detail=(
                "Current user is not associated "
                "with a company."
            ),
        )

    # --------------------------------------------------------
    # GET ANALYTICS
    # --------------------------------------------------------

    analytics_service = (
        OrganizationAnalyticsService(db)
    )

    overview = await (
        analytics_service
        .analytics_repository
        .get_overview(company_id)
    )

    # --------------------------------------------------------
    # CREATE WORKBOOK
    # --------------------------------------------------------

    workbook = Workbook()

    worksheet = workbook.active

    worksheet.title = "Overview"

    # --------------------------------------------------------
    # TITLE
    # --------------------------------------------------------

    worksheet["A1"] = "Recruitment Intelligence"
    worksheet["A1"].font = TITLE_FONT

    worksheet["A2"] = "Overview Report"
    worksheet["A2"].font = SUBTITLE_FONT

    worksheet["A3"] = (
        f"Company ID: {company_id}"
    )

    # --------------------------------------------------------
    # HEADERS
    # --------------------------------------------------------

    headers = [
        "Metric",
        "Value",
    ]

    header_row = 5

    for column_index, header in enumerate(
        headers,
        start=1,
    ):

        worksheet.cell(
            row=header_row,
            column=column_index,
            value=header,
        )

    style_header(
        worksheet,
        header_row,
    )

    # --------------------------------------------------------
    # DATA
    # --------------------------------------------------------

    metrics = [
        (
            "Total Jobs",
            overview.get(
                "total_jobs",
                0,
            ),
        ),
        (
            "Active Jobs",
            overview.get(
                "active_jobs",
                0,
            ),
        ),
        (
            "Applications",
            overview.get(
                "applications",
                0,
            ),
        ),
        (
            "Matched",
            overview.get(
                "matched",
                0,
            ),
        ),
        (
            "Shortlisted",
            overview.get(
                "shortlisted",
                0,
            ),
        ),
        (
            "Interviews",
            overview.get(
                "interviews",
                0,
            ),
        ),
        (
            "Finalists",
            overview.get(
                "finalists",
                0,
            ),
        ),
        (
            "Selected",
            overview.get(
                "selected",
                0,
            ),
        ),
        (
            "Rejected",
            overview.get(
                "rejected",
                0,
            ),
        ),
        (
            "Conversion (%)",
            overview.get(
                "conversion",
                0,
            ),
        ),
        (
            "Average ATS",
            overview.get(
                "avg_ats",
                0,
            ),
        ),
        (
            "Time to Hire (Days)",
            overview.get(
                "time_to_hire",
                0,
            ),
        ),
    ]

    data_row = header_row + 1

    for metric, value in metrics:

        worksheet.cell(
            row=data_row,
            column=1,
            value=metric,
        )

        worksheet.cell(
            row=data_row,
            column=2,
            value=value,
        )

        worksheet.cell(
            row=data_row,
            column=1,
        ).border = THIN_BORDER

        worksheet.cell(
            row=data_row,
            column=2,
        ).border = THIN_BORDER

        data_row += 1

    # --------------------------------------------------------
    # FORMATTING
    # --------------------------------------------------------

    conversion_row = header_row + 10

    worksheet.cell(
        row=conversion_row,
        column=2,
    ).number_format = "0.0"

    worksheet.freeze_panes = "A6"

    auto_width(worksheet)

    # --------------------------------------------------------
    # DOWNLOAD
    # --------------------------------------------------------

    return workbook_to_response(
        workbook,
        "overview_report.xlsx",
    )


# ============================================================
# 2. EXPORT JOB PERFORMANCE
# ============================================================

@router.get(
    "/export/job-performance",
)
async def export_job_performance(
    current_user=Depends(
        get_current_company_admin
    ),
    db: AsyncSession = Depends(get_db),
):

    # --------------------------------------------------------
    # COMPANY
    # --------------------------------------------------------

    company_id = current_user.company_id

    if company_id is None:
        from fastapi import HTTPException

        raise HTTPException(
            status_code=400,
            detail=(
                "Current user is not associated "
                "with a company."
            ),
        )

    # --------------------------------------------------------
    # GET ANALYTICS
    # --------------------------------------------------------

    analytics_service = (
        OrganizationAnalyticsService(db)
    )

    job_performance = await (
        analytics_service
        .analytics_repository
        .get_job_performance(company_id)
    )

    # --------------------------------------------------------
    # CREATE WORKBOOK
    # --------------------------------------------------------

    workbook = Workbook()

    worksheet = workbook.active

    worksheet.title = "Job Performance"

    # --------------------------------------------------------
    # TITLE
    # --------------------------------------------------------

    worksheet["A1"] = "Recruitment Intelligence"
    worksheet["A1"].font = TITLE_FONT

    worksheet["A2"] = "Job Performance Report"
    worksheet["A2"].font = SUBTITLE_FONT

    worksheet["A3"] = (
        f"Company ID: {company_id}"
    )

    # --------------------------------------------------------
    # HEADERS
    # --------------------------------------------------------

    headers = [
        
        "Job Title",
        "Department",
        "Applications",
        "Matched",
        "Match Rate (%)",
        "Shortlisted",
        "Interviews",
        "Selected",
        "Avg ATS",
        "Days Open",
        "Conversion (%)",
    ]

    header_row = 5

    for column_index, header in enumerate(
        headers,
        start=1,
    ):

        worksheet.cell(
            row=header_row,
            column=column_index,
            value=header,
        )

    style_header(
        worksheet,
        header_row,
    )

    # --------------------------------------------------------
    # JOB DATA
    # --------------------------------------------------------

    jobs = job_performance.get(
        "jobs",
        [],
    )

    row_number = header_row + 1

    for job in jobs:

        values = [
            
            job.get(
                "job_title",
                "",
            ),
            job.get(
                "department",
                "",
            ),
            job.get(
                "applications",
                0,
            ),
            job.get(
                "matched",
                0,
            ),
            job.get(
                "match_rate",
                0,
            ),
            job.get(
                "shortlisted",
                0,
            ),
            job.get(
                "interviews",
                0,
            ),
            job.get(
                "selected",
                0,
            ),
            job.get(
                "avg_ats",
                0,
            ),
            job.get(
                "days_open",
                0,
            ),
            job.get(
                "conversion",
                0,
            ),
        ]

        for column_index, value in enumerate(
            values,
            start=1,
        ):

            cell = worksheet.cell(
                row=row_number,
                column=column_index,
                value=value,
            )

            cell.border = THIN_BORDER

            cell.alignment = Alignment(
                vertical="center",
            )

        row_number += 1

    # --------------------------------------------------------
    # FORMATTING
    # --------------------------------------------------------

    worksheet.freeze_panes = "A6"

    worksheet.auto_filter.ref = (
    f"A5:K{max(row_number - 1, 5)}"
)

    auto_width(worksheet)

    # --------------------------------------------------------
    # DOWNLOAD
    # --------------------------------------------------------

    return workbook_to_response(
        workbook,
        "job_performance_report.xlsx",
    )